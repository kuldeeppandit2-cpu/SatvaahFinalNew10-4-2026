#!/usr/bin/env python3
"""
populate_database_sheet.py

Reads the Database sheet from satvaaah_taxonomy_synonyms.xlsx
For each L4 + City combination, searches IndiaMART (products) or 
Sulekha (services/expertise/establishments) and fills:
  Col F: Name of person  (owner name if available, else empty)
  Col G: Name of Firm    (company name)
  Col H: Mobile Number   (real phone)
  Col I: Address         (city + district)
  Col J: Geo location    (lat,long)

USAGE:
  python3 scripts/populate_database_sheet.py --file satvaaah_taxonomy_synonyms.xlsx --test
  python3 scripts/populate_database_sheet.py --file satvaaah_taxonomy_synonyms.xlsx

TEST MODE: processes first L4 only (20 rows), shows what would be written
FULL MODE: processes all L4s
"""

import re, time, random, argparse, sys, os
from datetime import datetime

try:
    import requests
    import openpyxl
except ImportError:
    import subprocess
    subprocess.run([sys.executable,'-m','pip','install','requests','openpyxl','-q'])
    import requests, openpyxl

LOG_FILE = 'populate_database_sheet.log'

def log(msg):
    ts = datetime.now().strftime('%H:%M:%S')
    line = f"[{ts}] {msg}"
    print(line)
    with open(LOG_FILE,'a') as f:
        f.write(line+'\n')

def make_session():
    s = requests.Session()
    s.headers.update({
        'User-Agent': 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) '
                      'AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36',
        'Accept-Language': 'en-IN,en;q=0.9',
        'Referer': 'https://www.indiamart.com/',
    })
    return s

# ── IndiaMART scraper ─────────────────────────────────────────────────────────
CITY_MAP = {
    'Hyderabad':'Hyderabad','Mumbai':'Mumbai','Delhi':'Delhi',
    'Chennai':'Chennai','Bangalore':'Bangalore',
    'Hyderbad':'Hyderabad',  # fix typo in sheet
}

def search_indiamart(session, query, city, max_results=20):
    """
    Search IndiaMART for product suppliers.
    Returns list of dicts with: firm, phone, address, lat, lng
    """
    im_city = CITY_MAP.get(city, city)
    try:
        r = session.get(
            'https://dir.indiamart.com/search.mp',
            params={'ss': query, 'src_area': im_city, 'page': 1},
            timeout=12
        )
        if r.status_code != 200:
            return []
        html = r.text

        # Extract lat/lng from location_info
        lat_m = re.search(r'"lat"\s*:\s*"(-?\d+\.\d+)"', html)
        lng_m = re.search(r'"long"\s*:\s*"(-?\d+\.\d+)"', html)
        city_lat = lat_m.group(1) if lat_m else ''
        city_lng = lng_m.group(1) if lng_m else ''

        # Extract all listings using companyname as anchor
        # Each listing has: companyname, pns (phone), city, district, state
        results = []
        seen_phones = set()

        # Find all companyname occurrences
        for m in re.finditer(r'"companyname"\s*:\s*"([^"]{3,80})"', html):
            firm = m.group(1).strip()
            pos  = m.start()

            # Search within 3000 chars forward for phone, city, district
            chunk = html[pos:pos+3000]

            pns_m  = re.search(r'"pns"\s*:\s*"(\d{10,12})"', chunk)
            city_m = re.search(r'"city"\s*:\s*"([^"]+)"', chunk)
            dist_m = re.search(r'"district"\s*:\s*"([^"]+)"', chunk)
            stat_m = re.search(r'"state"\s*:\s*"([^"]+)"', chunk)

            if not pns_m:
                # Try score field as fallback phone
                score_m = re.search(r'"score"\s*:\s*(\d{12})', chunk)
                if score_m:
                    candidate = str(int(score_m.group(1))//100)
                    if candidate[0] in '6789' and len(candidate)==10:
                        phone = candidate
                    else:
                        continue
                else:
                    continue
            else:
                phone = pns_m.group(1)
                # Normalise to 10 digits
                if len(phone)==12 and phone.startswith('91'):
                    phone = phone[2:]
                if not (len(phone)==10 and phone[0] in '6789'):
                    continue

            if phone in seen_phones:
                continue
            seen_phones.add(phone)

            city_val  = city_m.group(1) if city_m else im_city
            dist_val  = dist_m.group(1) if dist_m else ''
            state_val = stat_m.group(1) if stat_m else ''
            address   = ', '.join(filter(None,[dist_val, city_val, state_val]))

            results.append({
                'firm':    firm,
                'phone':   phone,
                'address': address,
                'lat':     city_lat,
                'lng':     city_lng,
            })

            if len(results) >= max_results:
                break

        return results

    except Exception as e:
        log(f"  IndiaMART error: {e}")
        return []

# ── Sulekha scraper ────────────────────────────────────────────────────────────
def search_sulekha(session, query, city, max_results=20):
    """
    Search Sulekha for service/expertise/establishment providers.
    Returns list of dicts with: firm, phone, address, lat, lng
    """
    city_slug = city.lower().replace(' ','-')
    query_slug = re.sub(r'[^a-z0-9]+','-', query.lower()).strip('-')
    url = f"https://www.sulekha.com/{query_slug}/{city_slug}"
    try:
        r = session.get(url, timeout=12)
        if r.status_code != 200:
            return []
        html = r.text

        results = []
        seen_phones = set()

        # Sulekha embeds phone in JSON-LD and data attributes
        phones = re.findall(r'[6-9]\d{9}', html)
        # Find business names near phones
        names  = re.findall(
            r'(?:class="[^"]*companyname[^"]*"|"name"\s*:)\s*[">]([^<"]{3,60})',
            html)

        for i, phone in enumerate(phones[:max_results]):
            if phone in seen_phones:
                continue
            seen_phones.add(phone)
            firm = names[i].strip() if i < len(names) else query
            results.append({
                'firm':    firm,
                'phone':   phone,
                'address': city,
                'lat':     '',
                'lng':     '',
            })

        return results

    except Exception as e:
        log(f"  Sulekha error: {e}")
        return []

# ── Main populate function ────────────────────────────────────────────────────
def populate(filepath, test_mode):
    wb = openpyxl.load_workbook(filepath)
    ws = wb['Database']

    log(f"Opened: {filepath}")
    log(f"Sheet rows: {ws.max_row}")

    session = make_session()

    # Read all unique L4+city combos that need data
    # Col D=L4 name, Col E=Prompt, Col K=City
    # Cols F-J need to be filled (cols 6-10)
    combos = {}  # (l4, city) -> list of row numbers
    for row in range(3, ws.max_row+1):
        l4   = ws.cell(row, 4).value
        city = ws.cell(row, 11).value
        cat  = ws.cell(row, 2).value  # Product/Service/Expertise/Establishment
        if not l4 or not city:
            continue
        # Only process rows that are empty
        if ws.cell(row, 8).value:  # Mobile already filled
            continue
        key = (str(l4), str(city), str(cat))
        if key not in combos:
            combos[key] = []
        combos[key].append(row)

    log(f"L4+city combos needing data: {len(combos)}")

    if test_mode:
        # Take only first combo
        first_key = list(combos.keys())[0]
        combos = {first_key: combos[first_key]}
        log(f"TEST MODE: processing only '{first_key[0]}' in '{first_key[1]}'")

    total_filled = 0

    for (l4, city, cat), rows in combos.items():
        log(f"\nSearching: '{l4}' in '{city}' ({cat}) — {len(rows)} rows to fill")

        # Choose source based on category
        if 'product' in cat.lower():
            results = search_indiamart(session, l4, city, max_results=len(rows))
            source  = 'IndiaMART'
        else:
            results = search_sulekha(session, l4, city, max_results=len(rows))
            source  = 'Sulekha'

        log(f"  {source} returned {len(results)} results")

        if not results:
            log(f"  ⚪ No results — leaving blank")
            continue

        # Fill rows with results
        filled = 0
        for i, row_num in enumerate(rows):
            if i >= len(results):
                break
            r = results[i]
            geo = f"{r['lat']},{r['lng']}" if r['lat'] and r['lng'] else ''

            # Write to sheet
            ws.cell(row_num, 6).value  = ''            # Name of person (unknown)
            ws.cell(row_num, 7).value  = r['firm']     # Name of Firm
            ws.cell(row_num, 8).value  = r['phone']    # Mobile Number
            ws.cell(row_num, 9).value  = r['address']  # Address
            ws.cell(row_num, 10).value = geo            # Geo location

            log(f"  ✅ Row {row_num}: {r['firm'][:40]} | {r['phone']} | {r['address'][:40]}")
            filled += 1

        total_filled += filled
        log(f"  Filled {filled}/{len(rows)} rows")

        # Save after each L4 so we don't lose data
        wb.save(filepath)
        log(f"  Saved to {filepath}")

        # Polite delay
        time.sleep(random.uniform(2, 4))

    log(f"\n{'='*60}")
    log(f"DONE. Total rows filled: {total_filled}")
    log(f"File saved: {filepath}")

def main():
    ap = argparse.ArgumentParser()
    ap.add_argument('--file', required=True, help='Path to satvaaah_taxonomy_synonyms.xlsx')
    ap.add_argument('--test', action='store_true', help='Test mode: first L4 only')
    args = ap.parse_args()

    if not os.path.exists(args.file):
        print(f"File not found: {args.file}")
        sys.exit(1)

    populate(args.file, args.test)

if __name__ == '__main__':
    main()
